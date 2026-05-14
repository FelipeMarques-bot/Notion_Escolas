from notion_client import Client
from notion_client.errors import RequestTimeoutError
import os
import time
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None

NOTION_TOKEN = os.environ.get("NOTION_TOKEN", "")
ROOT_PAGE_ID = os.environ.get("ROOT_PAGE_ID", "")

notion = Client(auth=NOTION_TOKEN)

ARQUIVO_ALUNOS_XLSX = os.environ.get("ARQUIVO_ALUNOS_XLSX", "Notas Escolas - 2° Trimestre.xlsx")
TRIMESTRE_IMPORTACAO_ALUNOS = os.environ.get("TRIMESTRE_IMPORTACAO_ALUNOS", "2º Trimestre")


def executar_notion_com_retry(funcao, tentativas=8, espera_base=1.0):
    ultima_excecao = None
    for tentativa in range(1, tentativas + 1):
        try:
            return funcao()
        except Exception as erro:
            mensagem = str(erro)
            msg_lower = mensagem.lower()
            recuperavel = (
                isinstance(erro, RequestTimeoutError)
                or "502" in mensagem
                or "timed out" in msg_lower
                or "rate limited" in msg_lower
                or "429" in mensagem
            )
            if not recuperavel:
                raise
            ultima_excecao = erro
            if tentativa == tentativas:
                raise
            # Backoff progressivo para reduzir falhas em burst na API do Notion.
            time.sleep(espera_base * tentativa * 2)
    raise ultima_excecao

ESCOLAS = [
    {"nome": "Juvenal",      "turnos": ["Matutino"],               "emoji": "🏫"},
    {"nome": "Arapongas",    "turnos": ["Vespertino"],             "emoji": "🏫"},
    {"nome": "Mulde",        "turnos": ["Matutino"],               "emoji": "🏫"},
    {"nome": "Anna Alves",   "turnos": ["Vespertino"],             "emoji": "🏫"},
    {"nome": "Tancredo",     "turnos": ["Matutino", "Vespertino"], "emoji": "🏫"},
    {"nome": "Maria Helena", "turnos": ["Matutino", "Vespertino"], "emoji": "🏫"},
]

TURMAS = ["6º Ano", "7º Ano", "8º Ano", "9º Ano"]
TRIMESTRES = ["1º Trimestre", "2º Trimestre", "3º Trimestre"]

KANBAN_STATUS_FLUXO = ["To-do", "In progress", "In review", "Complete"]

KANBAN_CARDS_REFERENCIA = {
    "To-do": [
        "Revisar pendências de lançamento",
        "Conferir alunos sem nota em Atividade 1",
        "Atualizar observações pedagógicas",
    ],
    "In progress": [
        "Fechamento parcial da turma",
        "Validação de status de entrega",
        "Conferência de médias provisórias",
    ],
    "In review": [
        "Auditoria de inconsistências",
        "Revisão final por coordenação",
        "Ajuste de planos de ação",
    ],
    "Complete": [
        "Trimestre fechado",
        "Relatório pedagógico atualizado",
        "Pendências resolvidas",
    ],
}

CAPAS_NIVEL = {
    "escola": "https://images.unsplash.com/photo-1523050854058-8df90110c9f1?auto=format&fit=crop&w=1400&q=80",
    "turno_matutino": "https://images.unsplash.com/photo-1509062522246-3755977927d7?auto=format&fit=crop&w=1400&q=80",
    "turno_vespertino": "https://images.unsplash.com/photo-1503676260728-1c00da094a0b?auto=format&fit=crop&w=1400&q=80",
    "turma": "https://images.unsplash.com/photo-1497486751825-1233686d5d80?auto=format&fit=crop&w=1400&q=80",
    "trimestre": "https://images.unsplash.com/photo-1434030216411-0b793f4b4173?auto=format&fit=crop&w=1400&q=80",
}

USAR_CAPAS_PADRAO = os.environ.get("USAR_CAPAS_PADRAO", "1") == "1"


def capa_nivel(chave):
    if USAR_CAPAS_PADRAO:
        return CAPAS_NIVEL.get(chave)
    return None

def icone_turno(turno):
    return {"Matutino": "🌅", "Vespertino": "🌤️"}.get(turno, "📚")

def icone_turma(turma):
    return {"6º Ano": "6️⃣", "7º Ano": "7️⃣", "8º Ano": "8️⃣", "9º Ano": "9️⃣"}.get(turma, "📚")

def icone_trimestre(trimestre):
    return {"1º Trimestre": "1️⃣", "2º Trimestre": "2️⃣", "3º Trimestre": "3️⃣"}.get(trimestre, "📋")


def bloco_link_pagina(page_id):
    return {
        "object": "block",
        "type": "link_to_page",
        "link_to_page": {"type": "page_id", "page_id": page_id},
    }


def bloco_toggle_menu(titulo, icone, page_ids):
    filhos = [bloco_link_pagina(page_id) for page_id in page_ids]
    return {
        "object": "block",
        "type": "toggle",
        "toggle": {
            "rich_text": [{"type": "text", "text": {"content": f"{icone} {titulo}"}}],
            "children": filhos,
            "color": "default",
        },
    }


def bloco_espaco_para_capa():
    return {
        "object": "block",
        "type": "callout",
        "callout": {
            "rich_text": [{"type": "text", "text": {"content": "🖼️ Espaço de capa personalizado: você pode trocar a capa desta página quando quiser."}}],
            "icon": {"type": "emoji", "emoji": "🎨"},
            "color": "gray_background",
        },
    }


def bloco_area_anotacoes(contexto):
    return [
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🗒️ Bloco de Anotações"}}]
        }},
        {"object": "block", "type": "toggle", "toggle": {
            "rich_text": [{"type": "text", "text": {"content": f"Registro rápido ({contexto})"}}],
            "children": [
                {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                    "rich_text": [{"type": "text", "text": {"content": "Pontos fortes da turma/escola:"}}]
                }},
                {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                    "rich_text": [{"type": "text", "text": {"content": "Alunos com atenção imediata:"}}]
                }},
                {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                    "rich_text": [{"type": "text", "text": {"content": "Ação combinada para a próxima aula:"}}]
                }},
            ],
            "color": "default",
        }},
    ]


def blocos_pagina_escola(nome, turnos, turnos_menu=None, solicitacao_db_id=None):
    turnos_str = " | ".join([f"{icone_turno(t)} {t}" for t in turnos])
    blocos = [
        bloco_espaco_para_capa(),
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": f"🏫 {nome}"}},
                          {"type": "text", "text": {"content": f"\nTurnos disponíveis: {turnos_str}"}},
                          {"type": "text", "text": {"content": "\nPainel principal da escola com atalhos para reduzir abertura de abas."}}],
            "icon": {"type": "emoji", "emoji": "🏫"},
            "color": "blue_background",
        }},
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🧭 Menu Inteligente da Escola"}}]
        }},
        {"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Use os atalhos abaixo para ir direto ao turno, depois selecionar turma e trimestre."}}]
        }},
    ]
    if turnos_menu:
        blocos.extend([bloco_link_pagina(pid) for pid in turnos_menu])
    else:
        blocos.append({"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Atalhos aparecerão automaticamente após a sincronização inicial desta escola."}}]
        }})

    blocos.extend([
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_3", "heading_3": {
            "rich_text": [{"type": "text", "text": {"content": "⚡ Lançamento no SGE (por escola)"}}]
        }},
        {"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Use o GitHub Actions (Run workflow) para solicitar o lançamento manualmente quando necessário."}}]
        }},
        {"object": "block", "type": "heading_3", "heading_3": {
            "rich_text": [{"type": "text", "text": {"content": "📌 Gestão de Avaliações"}}]
        }},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
            "rich_text": [{"type": "text", "text": {"content": "Padronize critérios de correção entre as turmas."}}]
        }},
        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
            "rich_text": [{"type": "text", "text": {"content": "Use os campos de status e observações para auditoria pedagógica."}}]
        }},
    ])

    if solicitacao_db_id:
        blocos.extend([
            {"object": "block", "type": "callout", "callout": {
                "rich_text": [{"type": "text", "text": {"content": "➡️ Use o workflow 'Lancar Notas SGE' no GitHub Actions para executar agora ou aguarde o agendamento automático."}}],
                "icon": {"type": "emoji", "emoji": "▶️"},
                "color": "yellow_background",
            }},
        ])

    blocos.extend(bloco_area_anotacoes(f"Escola {nome}"))
    return blocos


def blocos_pagina_turno(turno, nome_escola, turmas_menu=None):
    blocos = [
        bloco_espaco_para_capa(),
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": f"{nome_escola} — Turno {turno}"}},
                          {"type": "text", "text": {"content": "\nSelecione a turma abaixo no menu e acesse os trimestres sem navegar em múltiplas abas."}}],
            "icon": {"type": "emoji", "emoji": icone_turno(turno)},
            "color": "yellow_background",
        }},
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🎯 Menu de Turmas"}}]
        }},
    ]
    if turmas_menu:
        blocos.extend([bloco_link_pagina(pid) for pid in turmas_menu])
    else:
        blocos.append({"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "As turmas serão listadas automaticamente após a criação/sincronização."}}]
        }})

    blocos.extend(bloco_area_anotacoes(f"{nome_escola} - {turno}"))
    return blocos


def blocos_pagina_turma(turma, turno, nome_escola, trimestres_menu=None):
    blocos = [
        bloco_espaco_para_capa(),
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": f"{nome_escola} | {turno} | {turma}"}},
                          {"type": "text", "text": {"content": "\nCentro da turma com acesso rápido aos 3 trimestres e área de acompanhamento."}}],
            "icon": {"type": "emoji", "emoji": icone_turma(turma)},
            "color": "green_background",
        }},
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "📅 Menu de Trimestres"}}]
        }},
    ]
    if trimestres_menu:
        blocos.extend([bloco_link_pagina(pid) for pid in trimestres_menu])
    else:
        blocos.append({"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Os atalhos de trimestre serão preenchidos automaticamente ao final da sincronização."}}]
        }})

    blocos.extend([
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": "💡 Organização sugerida: mantenha nomes de atividades consistentes entre os trimestres."}}],
            "icon": {"type": "emoji", "emoji": "🧭"},
            "color": "brown_background",
        }},
    ])
    blocos.extend(bloco_area_anotacoes(f"{nome_escola} - {turno} - {turma}"))
    return blocos

def blocos_pagina_trimestre(trimestre, turma, turno, nome_escola):
    blocos = [
        bloco_espaco_para_capa(),
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": f"{nome_escola} | {turno} | {turma} | {trimestre}"}},
                          {"type": "text", "text": {"content": "\nÁrea de lançamento das 3 avaliações, controle de status e acompanhamento da média."}}],
            "icon": {"type": "emoji", "emoji": icone_trimestre(trimestre)},
            "color": "purple_background",
        }},
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "📝 Lançamento de Avaliações"}}]
        }},
        {"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Preencha abaixo as 3 avaliações do trimestre para cada aluno."}}]
        }},
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": "💡 Para renomear uma atividade, clique no nome da coluna no database abaixo e edite diretamente."}}],
            "icon": {"type": "emoji", "emoji": "💡"},
            "color": "gray_background",
        }},
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": "🧩 Para ficar no visual de quadro (como o da referência), mude a visualização do database para Board e agrupe por 'Status Fluxo'."}}],
            "icon": {"type": "emoji", "emoji": "🗂️"},
            "color": "blue_background",
        }},
    ]
    blocos.extend(bloco_area_anotacoes(f"{nome_escola} - {turno} - {turma} - {trimestre}"))
    return blocos


def criar_coluna_kanban(status, cards):
    cor_status = {
        "To-do": "gray_background",
        "In progress": "yellow_background",
        "In review": "blue_background",
        "Complete": "green_background",
    }.get(status, "gray_background")

    filhos = [
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": status}}],
            "icon": {"type": "emoji", "emoji": "📌"},
            "color": cor_status,
        }}
    ]

    for item in cards:
        filhos.append(
            {"object": "block", "type": "to_do", "to_do": {
                "rich_text": [{"type": "text", "text": {"content": item}}],
                "checked": False,
                "color": "default",
            }}
        )

    filhos.append(
        {"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "+ New page"}}]
        }}
    )

    return {"object": "block", "type": "column", "column": {"children": filhos}}


def bloco_kanban_estilo_notion():
    return [
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "📋 Quadro de Operação da Semana"}}]
        }},
        {"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Visual no estilo board para acompanhamento rápido da operação pedagógica."}}]
        }},
        {"object": "block", "type": "column_list", "column_list": {
            "children": [
                criar_coluna_kanban("To-do", KANBAN_CARDS_REFERENCIA["To-do"]),
                criar_coluna_kanban("In progress", KANBAN_CARDS_REFERENCIA["In progress"]),
                criar_coluna_kanban("In review", KANBAN_CARDS_REFERENCIA["In review"]),
                criar_coluna_kanban("Complete", KANBAN_CARDS_REFERENCIA["Complete"]),
            ]
        }},
    ]

def propriedades_database_alunos(nome_coluna_titulo="Nome do Aluno"):
    return {
        nome_coluna_titulo: {"title": {}},
        "Status Fluxo": {"select": {"options": [
            {"name": "To-do", "color": "default"},
            {"name": "In progress", "color": "yellow"},
            {"name": "In review", "color": "blue"},
            {"name": "Complete", "color": "green"},
        ]}},
        "Atividade 1": {"rich_text": {}},
        "Status lancamento 1": {"select": {"options": [
            {"name": "Lancada", "color": "green"},
        ]}},
        "Data realização 1": {"date": {}},
        "Observações 1": {"rich_text": {}},
        "Atividade 2": {"rich_text": {}},
        "Status lancamento 2": {"select": {"options": [
            {"name": "Lancada", "color": "green"},
        ]}},
        "Data realização 2": {"date": {}},
        "Observações 2": {"rich_text": {}},
        "Atividade 3": {"rich_text": {}},
        "Status lancamento 3": {"select": {"options": [
            {"name": "Lancada", "color": "green"},
        ]}},
        "Data realização 3": {"date": {}},
        "Observações 3": {"rich_text": {}},
        "Última Atualização": {"last_edited_time": {}},
    }

def listar_blocos_filhos(parent_id):
    todos = []
    cursor = None

    while True:
        if cursor:
            resposta = executar_notion_com_retry(
                lambda: notion.blocks.children.list(block_id=parent_id, start_cursor=cursor)
            )
        else:
            resposta = executar_notion_com_retry(
                lambda: notion.blocks.children.list(block_id=parent_id)
            )

        todos.extend(resposta.get("results", []))
        if not resposta.get("has_more"):
            break
        cursor = resposta.get("next_cursor")

    return todos


def normalizar_titulo_notion(texto):
    if not texto:
        return ""
    valor = str(texto).strip().lower().replace("°", "º")
    return re.sub(r"\s+", " ", valor)

def encontrar_pagina_filha(parent_id, titulo):
    for bloco in listar_blocos_filhos(parent_id):
        if bloco.get("type") == "child_page" and bloco.get("child_page", {}).get("title") == titulo and not bloco.get("archived", False):
            return bloco["id"]
    return None


def page_id_valido(page_id):
    try:
        executar_notion_com_retry(lambda: notion.pages.retrieve(page_id=page_id))
        return True
    except Exception:
        return False


def procurar_page_id_por_titulo_no_parent(parent_id, titulo):
    cursor = None
    while True:
        if cursor:
            resposta = executar_notion_com_retry(
                lambda: notion.search(
                    query=titulo,
                    filter={"property": "object", "value": "page"},
                    start_cursor=cursor,
                    page_size=100,
                )
            )
        else:
            resposta = executar_notion_com_retry(
                lambda: notion.search(
                    query=titulo,
                    filter={"property": "object", "value": "page"},
                    page_size=100,
                )
            )

        for page in resposta.get("results", []):
            if page.get("archived"):
                continue

            props = page.get("properties", {})
            titulo_page = ""
            if "title" in props and props["title"].get("type") == "title":
                titulo_page = "".join([t.get("plain_text", "") for t in props["title"].get("title", [])]).strip()
            else:
                for prop in props.values():
                    if prop.get("type") == "title":
                        titulo_page = "".join([t.get("plain_text", "") for t in prop.get("title", [])]).strip()
                        break

            parent = page.get("parent", {})
            if (
                titulo_page == titulo
                and parent.get("type") == "page_id"
                and parent.get("page_id") == parent_id
            ):
                return page.get("id")

        if not resposta.get("has_more"):
            break
        cursor = resposta.get("next_cursor")

    return None

def encontrar_database_filho(parent_id, titulo, titulos_alternativos=None):
    titulos_alvo = {normalizar_titulo_notion(titulo)}
    if titulos_alternativos:
        for titulo_alt in titulos_alternativos:
            titulos_alvo.add(normalizar_titulo_notion(titulo_alt))

    for bloco in listar_blocos_filhos(parent_id):
        if bloco.get("type") != "child_database" or bloco.get("archived", False):
            continue

        titulo_bloco = bloco.get("child_database", {}).get("title", "")
        if normalizar_titulo_notion(titulo_bloco) in titulos_alvo:
            return bloco["id"]
    return None


def listar_databases_filhos(parent_id):
    databases = []
    for bloco in listar_blocos_filhos(parent_id):
        if bloco.get("type") != "child_database" or bloco.get("archived", False):
            continue
        databases.append({
            "id": bloco.get("id"),
            "title": bloco.get("child_database", {}).get("title", ""),
        })
    return databases


def encontrar_pagina_em_parents(parent_ids, titulo):
    for parent_id in parent_ids:
        page_id = encontrar_pagina_filha(parent_id, titulo)
        if page_id:
            return page_id
    return None

def substituir_blocos_da_pagina(page_id, blocos):
    if not page_id_valido(page_id):
        print(f"⚠️ Page id inválido para atualização de blocos: {page_id}")
        return

    for bloco in listar_blocos_filhos(page_id):
        if bloco.get("archived", False):
            continue
        # Nao remove subpaginas/databases para nao apagar a hierarquia existente.
        if bloco.get("type") in {"child_page", "child_database"}:
            continue

        try:
            executar_notion_com_retry(lambda bloco_id=bloco["id"]: notion.blocks.delete(block_id=bloco_id))
        except Exception as erro:
            mensagem = str(erro).lower()
            if "archived" in mensagem:
                continue
            raise

    if blocos:
        executar_notion_com_retry(lambda: notion.blocks.children.append(block_id=page_id, children=blocos))

def criar_ou_atualizar_pagina(parent_id, titulo, emoji, blocos, cover_url=None, parent_busca_extra=None, titulos_alternativos=None):
    existente_id = encontrar_pagina_filha(parent_id, titulo)

    if existente_id and not page_id_valido(existente_id):
        existente_id = procurar_page_id_por_titulo_no_parent(parent_id, titulo)

    movida_para_parent = False

    if not existente_id and titulos_alternativos:
        for titulo_alt in titulos_alternativos:
            existente_id = encontrar_pagina_filha(parent_id, titulo_alt)
            if existente_id and not page_id_valido(existente_id):
                existente_id = procurar_page_id_por_titulo_no_parent(parent_id, titulo_alt)
            if existente_id:
                break

    if not existente_id and parent_busca_extra:
        existente_id = encontrar_pagina_em_parents(parent_busca_extra, titulo)
        if existente_id:
            executar_notion_com_retry(lambda: notion.pages.update(
                page_id=existente_id,
                parent={"type": "page_id", "page_id": parent_id},
            ))
            movida_para_parent = True

    if existente_id:
        payload_update = {
            "page_id": existente_id,
            "icon": {"type": "emoji", "emoji": emoji},
            "properties": {"title": {"title": [{"type": "text", "text": {"content": titulo}}]}},
        }
        if cover_url:
            payload_update["cover"] = {"type": "external", "external": {"url": cover_url}}

        executar_notion_com_retry(lambda: notion.pages.update(**payload_update))
        substituir_blocos_da_pagina(existente_id, blocos)
        return existente_id, not movida_para_parent

    payload_create = {
        "parent": {"page_id": parent_id},
        "icon": {"type": "emoji", "emoji": emoji},
        "properties": {"title": [{"type": "text", "text": {"content": titulo}}]},
        "children": blocos,
    }
    if cover_url:
        payload_create["cover"] = {"type": "external", "external": {"url": cover_url}}

    response = executar_notion_com_retry(lambda: notion.pages.create(**payload_create))
    return response["id"], True

def criar_ou_atualizar_database_alunos(parent_page_id, titulo, titulos_alternativos=None):
    existente_id = encontrar_database_filho(parent_page_id, titulo, titulos_alternativos=titulos_alternativos)

    if existente_id:
        executar_notion_com_retry(lambda: notion.databases.update(
            database_id=existente_id,
            title=[{"type": "text", "text": {"content": titulo}}],
            icon={"type": "emoji", "emoji": "📊"},
        ))
        data_source_id = obter_data_source_id(existente_id)
        if data_source_id:
            titulo_prop = obter_nome_coluna_titulo_data_source(data_source_id)
            executar_notion_com_retry(lambda: notion.data_sources.update(
                data_source_id=data_source_id,
                properties=propriedades_database_alunos(titulo_prop),
            ))
        return existente_id, False

    response = executar_notion_com_retry(lambda: notion.databases.create(
        parent={"type": "page_id", "page_id": parent_page_id},
        icon={"type": "emoji", "emoji": "📊"},
        title=[{"type": "text", "text": {"content": titulo}}],
    ))
    database_id = response["id"]
    data_source_id = obter_data_source_id(database_id)
    if data_source_id:
        titulo_prop = obter_nome_coluna_titulo_data_source(data_source_id)
        executar_notion_com_retry(lambda: notion.data_sources.update(
            data_source_id=data_source_id,
            properties=propriedades_database_alunos(titulo_prop),
        ))
    return database_id, True


def propriedades_database_solicitacao_escola(titulo_prop="Name"):
    return {
        titulo_prop: {
            "title": {},
        },
        "Escola": {
            "rich_text": {},
        },
        "Solicitar lancamento": {
            "checkbox": {},
        },
        "Status lancamento": {
            "select": {
                "options": [
                    {"name": "Pendente", "color": "yellow"},
                    {"name": "Em execucao", "color": "blue"},
                    {"name": "Concluido", "color": "green"},
                    {"name": "Erro", "color": "red"},
                ]
            },
        },
        "Ultima execucao": {
            "date": {},
        },
        "Data lancamento": {
            "date": {},
        },
        "Log execucao": {
            "rich_text": {},
        },
        "Dry run": {
            "checkbox": {},
        },
    }


def criar_ou_atualizar_database_solicitacao_escola(parent_page_id, nome_escola):
    titulo = f"Solicitacoes SGE - {nome_escola}"
    existente_id = encontrar_database_filho(parent_page_id, titulo)

    if existente_id:
        executar_notion_com_retry(lambda: notion.databases.update(
            database_id=existente_id,
            title=[{"type": "text", "text": {"content": titulo}}],
            icon={"type": "emoji", "emoji": "🚀"},
        ))
        data_source_id = obter_data_source_id(existente_id)
        if data_source_id:
            titulo_prop = obter_nome_coluna_titulo_data_source(data_source_id)
            executar_notion_com_retry(lambda: notion.data_sources.update(
                data_source_id=data_source_id,
                properties=propriedades_database_solicitacao_escola(titulo_prop),
            ))
        return existente_id, False

    response = executar_notion_com_retry(lambda: notion.databases.create(
        parent={"type": "page_id", "page_id": parent_page_id},
        icon={"type": "emoji", "emoji": "🚀"},
        title=[{"type": "text", "text": {"content": titulo}}],
    ))
    database_id = response["id"]
    data_source_id = obter_data_source_id(database_id)
    if data_source_id:
        titulo_prop = obter_nome_coluna_titulo_data_source(data_source_id)
        executar_notion_com_retry(lambda: notion.data_sources.update(
            data_source_id=data_source_id,
            properties=propriedades_database_solicitacao_escola(titulo_prop),
        ))
    return database_id, True


def garantir_item_solicitacao_escola(database_id, nome_escola):
    data_source_id = obter_data_source_id(database_id)
    if not data_source_id:
        return False

    data_source = executar_notion_com_retry(lambda: notion.data_sources.retrieve(data_source_id=data_source_id))
    propriedades = data_source.get("properties", {})
    titulo_prop = None
    for nome_prop, definicao in propriedades.items():
        if definicao.get("type") == "title":
            titulo_prop = nome_prop
            break
    if not titulo_prop:
        return False

    existente = False
    cursor = None
    while True:
        if cursor:
            resposta = executar_notion_com_retry(
                lambda: notion.data_sources.query(data_source_id=data_source_id, start_cursor=cursor)
            )
        else:
            resposta = executar_notion_com_retry(
                lambda: notion.data_sources.query(data_source_id=data_source_id)
            )

        for pagina in resposta.get("results", []):
            titulo_node = pagina.get("properties", {}).get(titulo_prop, {})
            texto = "".join([t.get("plain_text", "") for t in titulo_node.get("title", [])]).strip()
            if texto == "Solicitar lancamento da escola":
                existente = True
                break

        if existente or not resposta.get("has_more"):
            break
        cursor = resposta.get("next_cursor")

    if existente:
        return False

    props = {
        titulo_prop: {
            "title": [{"type": "text", "text": {"content": "Solicitar lancamento da escola"}}],
        },
        "Escola": {
            "rich_text": [{"type": "text", "text": {"content": nome_escola}}],
        },
        "Solicitar lancamento": {"checkbox": False},
        "Status lancamento": {"select": {"name": "Pendente"}},
        "Dry run": {"checkbox": False},
    }

    executar_notion_com_retry(lambda: notion.pages.create(
        parent={"data_source_id": data_source_id},
        properties=props,
    ))
    return True


def obter_data_source_id(database_id):
    database = executar_notion_com_retry(lambda: notion.databases.retrieve(database_id=database_id))
    data_sources = database.get("data_sources", [])
    if data_sources:
        return data_sources[0].get("id")
    return None


def obter_nome_coluna_titulo_data_source(data_source_id):
    if not data_source_id:
        return "Name"
    data_source = executar_notion_com_retry(lambda: notion.data_sources.retrieve(data_source_id=data_source_id))
    for nome_prop, definicao in data_source.get("properties", {}).items():
        if definicao.get("type") == "title":
            return nome_prop
    return "Name"


def normalizar_nome_aluno(nome):
    if nome is None:
        return None
    texto = str(nome).strip()
    if not texto:
        return None
    texto = re.sub(r"\s*\.\s*$", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    if texto.lower() in {"nome aluno", "none"}:
        return None
    return texto


def normalizar_rotulo_trimestre(texto):
    if not texto:
        return ""
    valor = str(texto).strip().lower()
    valor = valor.replace("°", "º")
    valor = re.sub(r"\s+", " ", valor)
    return valor


def mapear_contexto_planilha(nome_aba):
    base = nome_aba.strip().lower()
    regras = [
        ("juv", "Juvenal", "Matutino"),
        ("arap", "Arapongas", "Vespertino"),
        ("mulde", "Mulde", "Matutino"),
        ("anna", "Anna Alves", "Vespertino"),
        ("tancredo mat", "Tancredo", "Matutino"),
        ("tancredo ves", "Tancredo", "Vespertino"),
        ("maria mat", "Maria Helena", "Matutino"),
        ("maria ves", "Maria Helena", "Vespertino"),
    ]
    for chave, escola, turno in regras:
        if chave in base:
            return escola, turno
    return None, None


def carregar_alunos_da_planilha(caminho_arquivo):
    alunos_por_turma = {}

    if not openpyxl:
        print("⚠️ openpyxl não instalado; importação de alunos ignorada.")
        return alunos_por_turma

    if not os.path.exists(caminho_arquivo):
        print(f"⚠️ Arquivo de alunos não encontrado: {caminho_arquivo}")
        return alunos_por_turma

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    colunas_nome_por_turma = [1, 15, 29, 43]

    for aba in wb.worksheets:
        escola, turno = mapear_contexto_planilha(aba.title)
        if not escola or not turno:
            continue

        for indice_turma, coluna_nome in enumerate(colunas_nome_por_turma):
            if indice_turma >= len(TURMAS):
                break

            turma = TURMAS[indice_turma]
            chave = (escola, turno, turma)
            nomes_existentes = alunos_por_turma.setdefault(chave, [])
            nomes_set = set(nomes_existentes)

            for linha in range(2, aba.max_row + 1):
                nome = normalizar_nome_aluno(aba.cell(row=linha, column=coluna_nome).value)
                if nome and nome not in nomes_set:
                    nomes_existentes.append(nome)
                    nomes_set.add(nome)

    total = sum(len(v) for v in alunos_por_turma.values())
    print(f"👥 Alunos carregados da planilha: {total}")
    return alunos_por_turma


def listar_nomes_alunos_no_database(database_id):
    data_source_id = obter_data_source_id(database_id)
    if not data_source_id:
        return set(), None, {}

    data_source = executar_notion_com_retry(lambda: notion.data_sources.retrieve(data_source_id=data_source_id))
    propriedades = data_source.get("properties", {})

    titulo_prop = None
    for nome_prop, definicao in propriedades.items():
        if definicao.get("type") == "title":
            titulo_prop = nome_prop
            break

    if not titulo_prop:
        return set(), None, propriedades

    nomes = set()
    cursor = None
    while True:
        if cursor:
            resposta = executar_notion_com_retry(
                lambda: notion.data_sources.query(data_source_id=data_source_id, start_cursor=cursor)
            )
        else:
            resposta = executar_notion_com_retry(
                lambda: notion.data_sources.query(data_source_id=data_source_id)
            )

        for pagina in resposta.get("results", []):
            prop = pagina.get("properties", {}).get(titulo_prop, {})
            textos = prop.get("title", [])
            nome = "".join([t.get("plain_text", "") for t in textos]).strip()
            if nome:
                nomes.add(nome)

        if not resposta.get("has_more"):
            break
        cursor = resposta.get("next_cursor")

    return nomes, titulo_prop, propriedades


def popular_alunos_no_database(database_id, alunos):
    if not alunos:
        return 0

    data_source_id = obter_data_source_id(database_id)
    if not data_source_id:
        return 0

    nomes_existentes, titulo_prop, propriedades = listar_nomes_alunos_no_database(database_id)
    if not titulo_prop:
        return 0

    inseridos = 0
    for idx, aluno in enumerate(alunos, start=1):
        if aluno in nomes_existentes:
            continue

        props = {
            titulo_prop: {
                "title": [{"type": "text", "text": {"content": aluno}}]
            }
        }

        if "Número" in propriedades and propriedades["Número"].get("type") == "number":
            props["Número"] = {"number": idx}

        executar_notion_com_retry(lambda props=props: notion.pages.create(
            parent={"data_source_id": data_source_id},
            properties=props,
        ))
        nomes_existentes.add(aluno)
        inseridos += 1

    return inseridos


def reconciliar_alunos_databases_relacionados(parent_page_id, database_id_destino, titulo, titulos_alternativos=None):
    titulos_alvo = {normalizar_titulo_notion(titulo)}
    if titulos_alternativos:
        for titulo_alt in titulos_alternativos:
            titulos_alvo.add(normalizar_titulo_notion(titulo_alt))

    relacionados = []
    for db in listar_databases_filhos(parent_page_id):
        db_id = db.get("id")
        if not db_id or db_id == database_id_destino:
            continue

        titulo_db = normalizar_titulo_notion(db.get("title", ""))
        if titulo_db in titulos_alvo:
            relacionados.append(db_id)

    if not relacionados:
        return 0

    alunos_para_recuperar = []
    for db_id in relacionados:
        nomes_db, _, _ = listar_nomes_alunos_no_database(db_id)
        if nomes_db:
            alunos_para_recuperar.extend(sorted(nomes_db))

    if not alunos_para_recuperar:
        return 0

    # Insercao e deduplicacao ficam centralizadas nesta funcao.
    return popular_alunos_no_database(database_id_destino, alunos_para_recuperar)

def blocos_dashboard_raiz(estrutura_escolas=None):
    blocos = [
        bloco_espaco_para_capa(),
        {"object": "block", "type": "heading_1", "heading_1": {
            "rich_text": [{"type": "text", "text": {"content": "🎛️ Dashboard de Lançamentos"}}]
        }},
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": "Painel mestre para gestão de avaliações e notas, com atalhos inteligentes por escola para reduzir navegação em múltiplas abas."}}],
            "icon": {"type": "emoji", "emoji": "🚀"},
            "color": "blue_background",
        }},
        {"object": "block", "type": "divider", "divider": {}},
    ]

    blocos.extend(bloco_kanban_estilo_notion())

    blocos.extend([
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "column_list", "column_list": {
            "children": [
                {"object": "block", "type": "column", "column": {
                    "children": [
                        {"object": "block", "type": "heading_3", "heading_3": {
                            "rich_text": [{"type": "text", "text": {"content": "⚡ Rotina Rápida"}}]
                        }},
                        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Abra a escola no Menu Inteligente."}}]
                        }},
                        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Escolha turno e turma com os atalhos da própria página."}}]
                        }},
                        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Lance as 3 atividades no trimestre e valide a situação final."}}]
                        }},
                    ]
                }},
                {"object": "block", "type": "column", "column": {
                    "children": [
                        {"object": "block", "type": "heading_3", "heading_3": {
                            "rich_text": [{"type": "text", "text": {"content": "📌 Boas Práticas"}}]
                        }},
                        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Padronize nomes das atividades entre turmas."}}]
                        }},
                        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Revise status de entrega para evitar médias incorretas."}}]
                        }},
                        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Use observações para justificativas e pendências."}}]
                        }},
                        {"object": "block", "type": "bulleted_list_item", "bulleted_list_item": {
                            "rich_text": [{"type": "text", "text": {"content": "Acompanhe frequência e plano de ação para cada aluno."}}]
                        }},
                    ]
                }},
                {"object": "block", "type": "column", "column": {
                    "children": [
                        {"object": "block", "type": "heading_3", "heading_3": {
                            "rich_text": [{"type": "text", "text": {"content": "🎯 Resultado"}}]
                        }},
                        {"object": "block", "type": "callout", "callout": {
                            "rich_text": [{"type": "text", "text": {"content": "Interface organizada para operação diária, com menos cliques e mais visão de gestão."}}],
                            "icon": {"type": "emoji", "emoji": "✅"},
                            "color": "green_background",
                        }},
                    ]
                }},
            ]
        }},
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🏫 Escolas (Menu Inteligente)"}}]
        }},
    ])

    if estrutura_escolas:
        for escola in estrutura_escolas:
            blocos.extend([
                {"object": "block", "type": "divider", "divider": {}},
                {"object": "block", "type": "callout", "callout": {
                    "rich_text": [{"type": "text", "text": {"content": f"{escola['emoji']} {escola['nome']}"}}],
                    "icon": {"type": "emoji", "emoji": "🏫"},
                    "color": "gray_background",
                }},
                bloco_link_pagina(escola["page_id"]),
                bloco_toggle_menu("Turnos da escola", "🧭", [turno["page_id"] for turno in escola["turnos"]]),
            ])
    else:
        blocos.append({"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "Menu será preenchido automaticamente após a sincronização da estrutura."}}]
        }})

    blocos.extend([
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🗂️ Área de Coordenação"}}]
        }},
        {"object": "block", "type": "toggle", "toggle": {
            "rich_text": [{"type": "text", "text": {"content": "Pendências da semana"}}],
            "children": [
                {"object": "block", "type": "to_do", "to_do": {
                    "rich_text": [{"type": "text", "text": {"content": "Conferir turmas sem lançamento completo"}}],
                    "checked": False,
                    "color": "default",
                }},
                {"object": "block", "type": "to_do", "to_do": {
                    "rich_text": [{"type": "text", "text": {"content": "Revisar situações críticas (média < 5)"}}],
                    "checked": False,
                    "color": "default",
                }},
            ],
            "color": "default",
        }},
    ])

    return blocos


def blocos_portal_visual(estrutura_escolas=None):
    blocos = [
        bloco_espaco_para_capa(),
        {"object": "block", "type": "heading_1", "heading_1": {
            "rich_text": [{"type": "text", "text": {"content": "✨ Portal de Gestão de Avaliações"}}]
        }},
        {"object": "block", "type": "callout", "callout": {
            "rich_text": [{"type": "text", "text": {"content": "Painel principal moderno para operar lançamentos sem abrir várias abas. Comece pela escola e navegue pelos atalhos internos."}}],
            "icon": {"type": "emoji", "emoji": "🧠"},
            "color": "blue_background",
        }},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "📍 Dashboard de Lançamentos (Topo)"}}]
        }},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {
            "rich_text": [{"type": "text", "text": {"content": "Escolha a escola no menu abaixo."}}]
        }},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {
            "rich_text": [{"type": "text", "text": {"content": "Dentro da escola, use o menu inteligente para Turno -> Turma -> Trimestre."}}]
        }},
        {"object": "block", "type": "numbered_list_item", "numbered_list_item": {
            "rich_text": [{"type": "text", "text": {"content": "Registre notas e observações sem perder o contexto de gestão."}}]
        }},
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🏫 Escolas"}}]
        }},
    ]

    if estrutura_escolas:
        for escola in estrutura_escolas:
            blocos.extend([
                {"object": "block", "type": "divider", "divider": {}},
                {"object": "block", "type": "callout", "callout": {
                    "rich_text": [{"type": "text", "text": {"content": f"{escola['emoji']} {escola['nome']}"}}],
                    "icon": {"type": "emoji", "emoji": "🏫"},
                    "color": "gray_background",
                }},
                bloco_link_pagina(escola["page_id"]),
                bloco_toggle_menu("Navegação por turno", "🧭", [turno["page_id"] for turno in escola["turnos"]]),
            ])
    else:
        blocos.append({"object": "block", "type": "paragraph", "paragraph": {
            "rich_text": [{"type": "text", "text": {"content": "O menu das escolas será preenchido ao fim da sincronização."}}]
        }})

    blocos.extend([
        {"object": "block", "type": "divider", "divider": {}},
        {"object": "block", "type": "heading_2", "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": "🗒️ Blocos de Anotações de Gestão"}}]
        }},
        {"object": "block", "type": "toggle", "toggle": {
            "rich_text": [{"type": "text", "text": {"content": "Planejamento Pedagógico"}}],
            "children": [
                {"object": "block", "type": "to_do", "to_do": {
                    "rich_text": [{"type": "text", "text": {"content": "Turmas com mais pendências"}}],
                    "checked": False,
                    "color": "default",
                }},
                {"object": "block", "type": "to_do", "to_do": {
                    "rich_text": [{"type": "text", "text": {"content": "Planos de recuperação por escola"}}],
                    "checked": False,
                    "color": "default",
                }},
                {"object": "block", "type": "to_do", "to_do": {
                    "rich_text": [{"type": "text", "text": {"content": "Ações para melhoria de frequência"}}],
                    "checked": False,
                    "color": "default",
                }},
            ],
            "color": "default",
        }},
    ])

    return blocos


def criar_ou_atualizar_portal_visual(estrutura_escolas=None):
    portal_id, _ = criar_ou_atualizar_pagina(
        ROOT_PAGE_ID,
        "Portal de Gestão de Avaliações",
        "✨",
        blocos_portal_visual(estrutura_escolas),
        cover_url=capa_nivel("escola"),
    )
    return portal_id


def limpar_atalhos_portal_na_raiz(portal_id):
    blocos_raiz = listar_blocos_filhos(ROOT_PAGE_ID)
    for bloco in blocos_raiz:
        if bloco.get("type") == "link_to_page":
            link_info = bloco.get("link_to_page", {})
            if link_info.get("type") == "page_id" and link_info.get("page_id") == portal_id:
                executar_notion_com_retry(lambda bloco_id=bloco["id"]: notion.blocks.delete(block_id=bloco_id))
                continue

        if bloco.get("type") == "callout":
            textos = bloco.get("callout", {}).get("rich_text", [])
            conteudo = "".join(
                t.get("plain_text", "") for t in textos if isinstance(t, dict)
            )
            if "Novo layout disponível" in conteudo and "Portal de Gestão de Avaliações" in conteudo:
                executar_notion_com_retry(lambda bloco_id=bloco["id"]: notion.blocks.delete(block_id=bloco_id))

def criar_ou_atualizar_dashboard_raiz(estrutura_escolas=None):
    dashboard_id, criado = criar_ou_atualizar_pagina(
        ROOT_PAGE_ID,
        "Dashboard de Lançamentos",
        "🎛️",
        blocos_dashboard_raiz(estrutura_escolas),
        cover_url=capa_nivel("escola"),
    )
    acao = "criado" if criado else "atualizado"
    print(f"\n🎛️ Dashboard inicial {acao}.")
    return dashboard_id

def criar_estrutura_completa():
    portal_id = criar_ou_atualizar_portal_visual()
    criar_ou_atualizar_dashboard_raiz()
    alunos_por_turma = carregar_alunos_da_planilha(ARQUIVO_ALUNOS_XLSX)
    estrutura_escolas = []

    for escola in ESCOLAS:
        nome_escola = escola["nome"]
        turnos = escola["turnos"]
        emoji_escola = escola.get("emoji", "🏫")
        print(f"\n🏫 Sincronizando escola: {nome_escola}")

        escola_page_id, escola_criada = criar_ou_atualizar_pagina(
            portal_id,
            nome_escola,
            emoji_escola,
            blocos_pagina_escola(nome_escola, turnos),
            cover_url=capa_nivel("escola"),
        )
        print(f"   {'✅ Criada' if escola_criada else '♻️ Atualizada'}")

        solicitacao_db_id, _ = criar_ou_atualizar_database_solicitacao_escola(escola_page_id, nome_escola)
        item_solicitacao_criado = garantir_item_solicitacao_escola(solicitacao_db_id, nome_escola)
        if item_solicitacao_criado:
            print("   🚀 Painel de solicitacao criado")

        escola_menu = {"nome": nome_escola, "emoji": emoji_escola, "page_id": escola_page_id, "turnos": []}

        for turno in turnos:
            print(f"  {icone_turno(turno)} Turno: {turno}")
            cover_turno = capa_nivel("turno_matutino") if turno == "Matutino" else capa_nivel("turno_vespertino")
            turno_page_id, _ = criar_ou_atualizar_pagina(
                escola_page_id,
                f"{turno} — {nome_escola}",
                icone_turno(turno),
                blocos_pagina_turno(turno, nome_escola),
                cover_url=cover_turno,
            )
            turno_menu = {"nome": turno, "page_id": turno_page_id, "turmas": []}

            for turma in TURMAS:
                print(f"    {icone_turma(turma)} Turma: {turma}")
                turma_page_id, _ = criar_ou_atualizar_pagina(
                    turno_page_id,
                    turma,
                    icone_turma(turma),
                    blocos_pagina_turma(turma, turno, nome_escola),
                    cover_url=capa_nivel("turma"),
                )
                turma_menu = {"nome": turma, "page_id": turma_page_id, "trimestres": []}

                for trimestre in TRIMESTRES:
                    print(f"      {icone_trimestre(trimestre)} {trimestre}")
                    trimestre_page_id, _ = criar_ou_atualizar_pagina(
                        turma_page_id,
                        trimestre,
                        icone_trimestre(trimestre),
                        blocos_pagina_trimestre(trimestre, turma, turno, nome_escola),
                        cover_url=capa_nivel("trimestre"),
                        titulos_alternativos=[trimestre.replace("º", "°")],
                    )
                    titulo_database = f"Notas Escolas - {trimestre} | {nome_escola} | {turno} | {turma}"
                    titulos_alt_database = [
                        f"Notas Escolas - {trimestre.replace('º', '°')} | {nome_escola} | {turno} | {turma}"
                    ]

                    database_id, _ = criar_ou_atualizar_database_alunos(
                        trimestre_page_id,
                        titulo_database,
                        titulos_alternativos=titulos_alt_database,
                    )

                    qtd_recuperados_notion = reconciliar_alunos_databases_relacionados(
                        trimestre_page_id,
                        database_id,
                        titulo_database,
                        titulos_alternativos=titulos_alt_database,
                    )
                    if qtd_recuperados_notion:
                        print(f"        ♻️ {qtd_recuperados_notion} alunos recuperados de databases anteriores")

                    if normalizar_rotulo_trimestre(trimestre) == normalizar_rotulo_trimestre(TRIMESTRE_IMPORTACAO_ALUNOS):
                        chave = (nome_escola, turno, turma)
                        qtd_inseridos = popular_alunos_no_database(
                            database_id,
                            alunos_por_turma.get(chave, []),
                        )
                        if qtd_inseridos:
                            print(f"        👥 {qtd_inseridos} alunos inseridos")
                    turma_menu["trimestres"].append({"nome": trimestre, "page_id": trimestre_page_id})

                substituir_blocos_da_pagina(
                    turma_page_id,
                    blocos_pagina_turma(
                        turma,
                        turno,
                        nome_escola,
                        trimestres_menu=[item["page_id"] for item in turma_menu["trimestres"]],
                    ),
                )
                turno_menu["turmas"].append(turma_menu)

            substituir_blocos_da_pagina(
                turno_page_id,
                blocos_pagina_turno(
                    turno,
                    nome_escola,
                    turmas_menu=[item["page_id"] for item in turno_menu["turmas"]],
                ),
            )
            escola_menu["turnos"].append(turno_menu)

        substituir_blocos_da_pagina(
            escola_page_id,
            blocos_pagina_escola(
                nome_escola,
                turnos,
                turnos_menu=[item["page_id"] for item in escola_menu["turnos"]],
                solicitacao_db_id=solicitacao_db_id,
            ),
        )
        estrutura_escolas.append(escola_menu)

    criar_ou_atualizar_dashboard_raiz(estrutura_escolas)
    criar_ou_atualizar_portal_visual(estrutura_escolas)
    limpar_atalhos_portal_na_raiz(portal_id)

    print("\n🎉 Estrutura completa sincronizada com sucesso no Notion!")

if __name__ == "__main__":
    criar_estrutura_completa()
